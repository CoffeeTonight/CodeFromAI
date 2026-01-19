from openpyxl import load_workbook
import pandas as pd
from pathlib import Path


def excel_to_html_table(ws):
    """Copy merged cell values, keep genuine empty cells empty"""
    data = [[cell.value for cell in row] for row in ws.iter_rows()]
    df = pd.DataFrame(data)

    # Copy values from merged cells
    for merged in ws.merged_cells.ranges:
        value = ws.cell(merged.min_row, merged.min_col).value
        for r in range(merged.min_row - 1, merged.max_row):
            for c in range(merged.min_col - 1, merged.max_col):
                if r < len(df) and c < len(df.columns):
                    df.iloc[r, c] = value

    df_display = df.astype(object).where(df.notnull(), "")

    html_table = df_display.to_html(
        index=False,
        header=False,
        border=1,
        na_rep="",
        escape=False,
        classes="excel-table"
    )

    full_html = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>{ws.title} - Excel Table</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; }}
        h2 {{ color: #333; margin-bottom: 15px; }}
        .excel-table {{
            border-collapse: collapse;
            width: auto;
            max-width: 100%;
            margin: 20px 0;
            font-size: 10pt;
        }}
        .excel-table td {{
            border: 1px solid #ccc;
            padding: 6px 10px;
            text-align: left;
            vertical-align: top;
            white-space: pre-wrap;
            min-width: 50px;
        }}
        .excel-table tr:nth-child(even) {{ background-color: #f9f9f9; }}
    </style>
</head>
<body>
    <h2>Sheet: {ws.title}</h2>
    {html_table}
</body>
</html>"""

    return full_html


def xlsx_to_html_to_md(
    input_xlsx="memorymap.xlsx",
    html_dir="html_temp",
    md_dir="markdown_output"
):
    """xlsx → HTML → Markdown conversion pipeline"""
    wb = load_workbook(input_xlsx, data_only=True)

    Path(html_dir).mkdir(parents=True, exist_ok=True)
    Path(md_dir).mkdir(parents=True, exist_ok=True)

    print(f"Processing file: {input_xlsx}")
    print(f"Number of sheets: {len(wb.sheetnames)}\n")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        html_content = excel_to_html_table(ws)

        safe_name = "".join(c if c.isalnum() or c in ('_', '-') else '_' for c in sheet_name)
        html_path = Path(html_dir) / f"{safe_name}.html"
        html_path.write_text(html_content, encoding="utf-8")
        print(f"HTML created: {html_path}")

        # f-string 들여쓰기 문제를 피하기 위해 괄호 + dedent 방식 사용
        md_content = (
            f"# Sheet: {sheet_name}\n\n"
            "This file preserves the original Excel sheet structure as much as possible.\n"
            "Due to complex merged cells (rowspan/colspan), perfect conversion to Markdown table is difficult.\n\n"
            "The original HTML table is included below.\n\n"
            "```html\n"
            f"{html_content}\n"
            "```\n\n"
            "**Notes**\n"
            f"- Open the original HTML file ({html_path.name}) in a browser for visual confirmation.\n"
            "- Merged cell values are copied across the merged range.\n"
            "- Genuine empty cells remain empty.\n"
        )

        md_path = Path(md_dir) / f"{safe_name}.md"
        md_path.write_text(md_content, encoding="utf-8")
        print(f"Markdown created: {md_path}\n")

    print("=== Conversion completed ===")
    print(f"HTML files: {html_dir}/")
    print(f"Markdown files: {md_dir}/")


if __name__ == "__main__":
    try:
        xlsx_to_html_to_md()
    except Exception as e:
        print("Error:", str(e))