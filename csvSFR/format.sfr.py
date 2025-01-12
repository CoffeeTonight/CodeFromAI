from openpyxl import Workbook, load_workbook

class ExcelMemoryMap:
    def __init__(self, excel_file):
        self.excel_file = excel_file
        self.workbook = load_workbook(filename=self.excel_file)
        self.summary_data = []

    def gather_sheet_info(self):
        for sheet in self.workbook.sheetnames:
            worksheet = self.workbook[sheet]
            first_cell_address = worksheet.cell(row=1, column=1).coordinate
            is_active = worksheet.sheet_state == 'visible'
            # Assuming each sheet represents a different AMBA bus interface
            bus_interface = "AMBA Bus"  # Placeholder for actual bus interface info
            if "AHB" in sheet:
                bus_interface = "AHB"
            elif "APB" in sheet:
                bus_interface = "APB"
            elif "AXI" in sheet:
                bus_interface = "AXI"
            
            self.summary_data.append({
                "Sheet Name": sheet,
                "First Address": first_cell_address,
                "Active": is_active,
                "Bus Interface": bus_interface
            })

    def add_summary_sheet(self):
        if 'Memory Map' not in self.workbook.sheetnames:
            summary_sheet = self.workbook.create_sheet(title='Memory Map')
        else:
            summary_sheet = self.workbook['Memory Map']
        
        headers = ["Sheet Name", "First Address", "Active", "Bus Interface"]
        summary_sheet.append(headers)
        
        for data in self.summary_data:
            summary_sheet.append([data["Sheet Name"], data["First Address"], data["Active"], data["Bus Interface"]])

    def save_workbook(self):
        self.workbook.save(filename=self.excel_file)

if __name__ == "__main__":
    excel_file = 'your_excel_file.xlsx'
    memory_map = ExcelMemoryMap(excel_file)
    memory_map.gather_sheet_info()
    memory_map.add_summary_sheet()
    memory_map.save_workbook()